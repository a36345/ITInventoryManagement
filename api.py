"""
api.py — REST API read-only para a interface web.
Corre no servidor ao lado da app desktop.
Inicia automaticamente quando a app desktop arranca (ver main.py).
Porta padrão: 5050
"""

from flask import Flask, jsonify, request, send_file, abort
from flask_cors import CORS
from datetime import datetime, timedelta
import io
import os

import core.database as db

app = Flask(__name__)
_cors_applied = False

_WEB_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "web")

API_KEY = None  # definido em runtime via set_api_key()


def _apply_cors():
    global _cors_applied
    if _cors_applied:
        return
    raw = db.get_setting(
        "web_cors_origins",
        "http://localhost:5050,http://127.0.0.1:5050,null",
    )
    origins = [o.strip() for o in raw.split(",") if o.strip()]
    if not origins:
        origins = ["http://localhost:5050", "null"]
    CORS(app, origins=origins, supports_credentials=True)
    _cors_applied = True

def set_api_key(key: str):
    global API_KEY
    API_KEY = key

def _auth():
    """Opcional: verifica header X-API-Key se configurado."""
    if not API_KEY:
        return True
    return request.headers.get("X-API-Key") == API_KEY

def _require_auth():
    if not _auth():
        abort(401, description="API Key inválida ou em falta.")

# ── Web UI ───────────────────────────────────────────────────────────────────

@app.get("/")
@app.get("/index.html")
def serve_ui():
    """Serve the web UI directly so the browser always loads the latest version."""
    return send_file(os.path.join(_WEB_DIR, "index.html"))

# ── Dashboard / Stats ────────────────────────────────────────────────────────

@app.get("/api/stats")
def get_stats():
    _require_auth()
    stats = db.get_stats()
    lifecycle = [dict(r) for r in db.get_lifecycle_report()]
    alerts = [dict(r) for r in db.get_open_alerts()[:10]]
    return jsonify({
        "stats":     stats,
        "lifecycle": lifecycle,
        "alerts":    alerts,
        "generated": datetime.utcnow().isoformat(),
    })

# ── Assets ────────────────────────────────────────────────────────────────────

@app.get("/api/assets")
def get_assets():
    _require_auth()
    type_    = request.args.get("type")
    dept     = request.args.get("department")
    status   = request.args.get("status")
    search   = request.args.get("search")
    review   = request.args.get("needs_review")
    assets   = db.get_all_assets(
        filter_type=type_, filter_dept=dept,
        filter_status=status, search=search,
        needs_review=True if review == "1" else None)
    return jsonify([dict(a) for a in assets])

@app.get("/api/assets/<int:asset_id>")
def get_asset(asset_id):
    _require_auth()
    a = db.get_asset(asset_id)
    if not a: abort(404)
    return jsonify(dict(a))

@app.delete("/api/assets/all")
def delete_all_assets():
    _require_auth()
    deleted = db.delete_all_assets()
    return jsonify({"ok": True, "deleted": deleted})

@app.delete("/api/assets/<int:asset_id>")
def delete_asset(asset_id):
    _require_auth()
    a = db.get_asset(asset_id)
    if not a:
        abort(404, description="Ativo não encontrado.")
    db.delete_asset(asset_id)
    db.invalidate_caches()
    return jsonify({"ok": True, "id": asset_id})

@app.get("/api/assets/<int:asset_id>/uptime")
def get_asset_uptime(asset_id):
    _require_auth()
    uptime = db.get_uptime_24h(asset_id)
    with db.get_conn() as c:
        history = c.execute("""
            SELECT is_online, ping_ms, checked_at
            FROM device_history WHERE asset_id=?
            AND checked_at >= datetime('now','-24 hours')
            ORDER BY checked_at
        """, (asset_id,)).fetchall()
    return jsonify({
        "uptime_pct": uptime,
        "history": [dict(h) for h in history]
    })

@app.get("/api/assets/stats/by-type")
def assets_by_type():
    _require_auth()
    with db.get_conn() as c:
        rows = c.execute(
            "SELECT type, COUNT(*) as count FROM assets GROUP BY type ORDER BY count DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.get("/api/assets/stats/by-department")
def assets_by_dept():
    _require_auth()
    with db.get_conn() as c:
        rows = c.execute(
            "SELECT department, COUNT(*) as count FROM assets WHERE department IS NOT NULL GROUP BY department ORDER BY count DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

# ── Printers ──────────────────────────────────────────────────────────────────

@app.get("/api/printers")
def get_printers():
    _require_auth()
    assets = db.get_all_assets(filter_type="Impressora")
    return jsonify([dict(a) for a in assets])

@app.get("/api/printers/critical")
def get_critical_printers():
    _require_auth()
    printers = db.get_critical_printers()
    return jsonify([dict(p) for p in printers])

# ── Alerts ────────────────────────────────────────────────────────────────────

@app.get("/api/alerts")
def get_alerts():
    _require_auth()
    severity = request.args.get("severity")
    alerts   = db.get_open_alerts()
    if severity:
        alerts = [a for a in alerts if a["severity"] == severity]
    return jsonify([dict(a) for a in alerts])

# ── Consumables ───────────────────────────────────────────────────────────────

@app.get("/api/consumables")
def get_consumables():
    _require_auth()
    return jsonify([dict(c) for c in db.get_all_consumables()])

@app.get("/api/consumables/low-stock")
def get_low_stock():
    _require_auth()
    return jsonify([dict(c) for c in db.get_low_stock_consumables()])

# ── Lifecycle ─────────────────────────────────────────────────────────────────

@app.get("/api/lifecycle")
def get_lifecycle():
    _require_auth()
    rows = db.get_lifecycle_report()
    current_year = datetime.now().year
    return jsonify([{
        "year":         r["acquisition_year"],
        "count":        r["cnt"],
        "age":          r["age"],
        "replace_due":  r["age"] >= 4,
    } for r in rows])

# ── Reports ───────────────────────────────────────────────────────────────────

@app.get("/api/reports/inventory.xlsx")
def report_xlsx():
    _require_auth()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        abort(500, "openpyxl não instalado.")

    wb = Workbook()

    # Sheet 1 — All assets
    ws = wb.active
    ws.title = "Inventário Completo"
    headers = ["ID","Hostname","Tipo","Estado","IP","MAC","Fabricante","Modelo",
               "S/N","Departamento","Utilizador","OS","Ano Aquis.","Preço €",
               "Fornecedor","Garantia","Última vez visto"]
    header_fill = PatternFill("solid", fgColor="1e2535")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    assets = db.get_all_assets()
    for row_idx, a in enumerate(assets, 2):
        values = [a["id"], a["hostname"], a["type"], a["status"],
                  a["ip_address"], a["mac_address"], a["manufacturer"],
                  a["model"], a["serial_number"], a["department"],
                  a["assigned_user"], a["os_version"], a["acquisition_year"],
                  a["purchase_price"], a["supplier"], a["warranty_years"],
                  (a["last_seen"] or "")[:16]]
        for col, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=v)
        if a["status"] == "Offline":
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="2d1515")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet 2 — Replacement due
    ws2 = wb.create_sheet("Substituição Planeada")
    ws2.append(["Hostname","Tipo","Departamento","Utilizador","Ano Aquis.","Anos em uso","IP"])
    ws2["A1"].font = Font(bold=True)
    current_year = datetime.now().year
    for a in assets:
        if a["type"] in ("Desktop","Laptop") and a["acquisition_year"]:
            age = current_year - a["acquisition_year"]
            if age >= 4:
                ws2.append([a["hostname"], a["type"], a["department"],
                            a["assigned_user"], a["acquisition_year"], age,
                            a["ip_address"]])

    # Sheet 3 — Printers
    ws3 = wb.create_sheet("Impressoras SNMP")
    ws3.append(["Hostname","IP","Toner K%","Toner C%","Toner M%","Toner Y%",
                "Total páginas","Último poll"])
    ws3["A1"].font = Font(bold=True)
    printers = db.get_all_assets(filter_type="Impressora")
    for p in printers:
        ws3.append([p["hostname"], p["ip_address"],
                    p["toner_black"], p["toner_cyan"],
                    p["toner_magenta"], p["toner_yellow"],
                    p["total_pages"],
                    (p["last_poll"] or "")[:16]])

    # Sheet 4 — Consumables
    ws4 = wb.create_sheet("Consumíveis")
    ws4.append(["Referência","Tipo","Compatível com","Stock","Mínimo","Estado"])
    ws4["A1"].font = Font(bold=True)
    for c in db.get_all_consumables():
        estado = "Sem stock" if c["stock_qty"]==0 else "Baixo" if c["stock_qty"]<c["stock_min"] else "OK"
        ws4.append([c["reference"], c["type"], c["compatible_with"],
                    c["stock_qty"], c["stock_min"], estado])

    # Sheet 5 — Open alerts
    ws5 = wb.create_sheet("Alertas")
    ws5.append(["Prioridade","Tipo","Título","Equipamento","Data","Estado"])
    ws5["A1"].font = Font(bold=True)
    for al in db.get_open_alerts():
        ws5.append([al["severity"], al["type"], al["title"],
                    al["hostname"], al["created_at"], al["status"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"IT_Inventory_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)

@app.get("/api/reports/summary.json")
def report_json():
    _require_auth()
    stats    = db.get_stats()
    assets   = [dict(a) for a in db.get_all_assets()]
    alerts   = [dict(a) for a in db.get_open_alerts()]
    consumes = [dict(c) for c in db.get_all_consumables()]
    return jsonify({
        "generated": datetime.utcnow().isoformat(),
        "stats": stats,
        "assets": assets,
        "alerts": alerts,
        "consumables": consumes,
        "lifecycle": [dict(r) for r in db.get_lifecycle_report()],
    })

# ── Network monitoring ────────────────────────────────────────────────────────

@app.get("/api/network/status")
def network_status():
    _require_auth()
    from core.network_monitor import get_live_status
    return jsonify(get_live_status())

@app.get("/api/network/history")
def network_history():
    _require_auth()
    hours = min(168, max(1, int(request.args.get("hours", 24))))
    metrics = [dict(r) for r in db.get_network_metrics_history(hours)]
    pings = {}
    for key in ("gateway", "firewall"):
        pings[key] = [dict(r) for r in db.get_network_ping_history(key, hours)]
    return jsonify({
        "hours": hours,
        "bandwidth": metrics,
        "pings": pings,
        "generated": datetime.utcnow().isoformat(),
    })

# ── Health ────────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    stats = db.get_stats()
    return jsonify({"status":"ok","assets":stats["total"],"ts":datetime.utcnow().isoformat()})

def start_api(host="0.0.0.0", port=5050, api_key=None, debug=False):
    db.init_db()
    _apply_cors()
    if api_key:
        set_api_key(api_key)
    import threading
    t = threading.Thread(
        target=lambda: app.run(host=host, port=port, debug=debug, use_reloader=False),
        daemon=True)
    t.start()
    return t
